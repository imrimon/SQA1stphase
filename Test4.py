import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import datetime
import time

# Load Excel file
excel_file = "Excel.xlsx"
wb = openpyxl.load_workbook(excel_file)

# Get today's day name
today = datetime.now().strftime('%A')

# Select the sheet corresponding to today's day name
today_sheet = wb[today]

# Read data from Excel
keywords_column = today_sheet['C']

# Initialize WebDriver
driver = webdriver.Chrome()  # You can change to your preferred WebDriver

# Iterate through the keywords in the current day's sheet
for keyword_row in keywords_column[1:]:
    keyword = keyword_row.value

    if keyword:  # Check if the cell is not empty
        # Open Google and interact with the search box
        driver.get("https://www.google.com")
        search_box = driver.find_element("name", "q")
        search_box.clear()
        search_box.send_keys(keyword)

        # Wait for search suggestions to appear (explicit wait)
        #time.sleep(1)  # Add a 2-second delay
        suggestions = driver.find_elements("xpath", "//ul/li[@role='presentation']")

        # Calculate the character count for each suggestion and choose shortest and longest
        suggestion_texts = [suggestion.text for suggestion in suggestions]
        shortest_suggestion = min(suggestion_texts, key=len) if suggestions else "No suggestions found"
        longest_suggestion = max(suggestion_texts, key=len) if suggestions else "No suggestions found"

        # Write suggestions in respective columns
        shortest_suggestion_col = today_sheet.max_column + 1
        longest_suggestion_col = today_sheet.max_column + 1

        keyword_row.offset(column=shortest_suggestion_col).value = shortest_suggestion
        keyword_row.offset(column=longest_suggestion_col).value = longest_suggestion

# Save Excel file with updated results
wb.save("excel_file14.xlsx")

# Close the WebDriver
driver.quit()
